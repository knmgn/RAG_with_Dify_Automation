"""
経費精算テンプレート_v3.xlsx ビルドスクリプト

株式会社デモ・ロジスティクス 旅費交通費・経費精算規定（DL-HR-RG-2024-007）に
準拠したダミーの経費精算 Excel テンプレートを生成する。

Usage:
    python3 scripts/build_expense_template.py

Output:
    ./経費精算テンプレート_v3.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.dimensions import ColumnDimension


OUTPUT_FILENAME = "経費精算テンプレート_v3.xlsx"

NAVY = "1F3864"
LIGHT_NAVY = "2E5A9E"
GREY_BG = "F2F2F2"
ALT_ROW = "FAFAFA"
WARN_BG = "FFF2CC"
ACCENT = "C00000"

THIN = Side(style="thin", color="999999")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title_font(size: int = 14, color: str = "FFFFFF") -> Font:
    return Font(name="Yu Gothic", size=size, bold=True, color=color)


def _body_font(size: int = 10, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Yu Gothic", size=size, bold=bold, color=color)


def _set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_main_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "経費精算書"
    ws.sheet_view.showGridLines = False

    _set_col_widths(
        ws,
        {
            "A": 4,
            "B": 12,
            "C": 14,
            "D": 28,
            "E": 26,
            "F": 18,
            "G": 14,
            "H": 10,
            "I": 24,
        },
    )

    ws.merge_cells("A1:I1")
    ws["A1"] = "経費精算書（Expense Reimbursement Form）"
    ws["A1"].font = _title_font(16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    ws["A2"] = "様式番号: F-001-v3 ／ 準拠規程: DL-HR-RG-2024-007（第3版・2024年10月1日改訂）"
    ws["A2"].font = _body_font(9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    header_rows = [
        ("A4", "申請日",     "B4", "",                    "D4", "精算番号",      "E4", ""),
        ("A5", "氏名",       "B5", "",                    "D5", "社員番号",      "E5", ""),
        ("A6", "所属部門",   "B6", "",                    "D6", "役職",          "E6", ""),
        ("A7", "対象期間",   "B7", "",                    "D7", "支払希望日",    "E7", ""),
    ]
    for label_cell, label, value_cell, _, label2_cell, label2, value2_cell, _ in header_rows:
        ws[label_cell] = label
        ws[label_cell].font = _body_font(10, bold=True)
        ws[label_cell].fill = PatternFill("solid", fgColor=GREY_BG)
        ws[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label_cell].border = BORDER_ALL

        ws.merge_cells(f"{value_cell}:C{value_cell[1:]}")
        ws[value_cell].border = BORDER_ALL
        ws[value_cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)

        ws[label2_cell] = label2
        ws[label2_cell].font = _body_font(10, bold=True)
        ws[label2_cell].fill = PatternFill("solid", fgColor=GREY_BG)
        ws[label2_cell].alignment = Alignment(horizontal="center", vertical="center")
        ws[label2_cell].border = BORDER_ALL

        ws.merge_cells(f"{value2_cell}:I{value2_cell[1:]}")
        ws[value2_cell].border = BORDER_ALL
        ws[value2_cell].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws["B4"] = "2026/05/26"
    ws["E4"] = "EX-2026-05-0123"
    ws["B5"] = "山田 太郎"
    ws["E5"] = "DL-04321"
    ws["B6"] = "営業本部 第一営業部"
    ws["E6"] = "主任"
    ws["B7"] = "2026/05/01 ～ 2026/05/25"
    ws["E7"] = "2026/06/25（給与振込日）"

    detail_header_row = 9
    detail_headers = [
        ("A", "No."),
        ("B", "利用日"),
        ("C", "区分"),
        ("D", "件名・目的"),
        ("E", "利用区間／場所"),
        ("F", "同行者・参加者"),
        ("G", "金額（税込）"),
        ("H", "領収書"),
        ("I", "備考（F-021等）"),
    ]
    for col, label in detail_headers:
        cell = ws[f"{col}{detail_header_row}"]
        cell.value = label
        cell.font = _title_font(10)
        cell.fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[detail_header_row].height = 28

    sample_rows = [
        (1, "2026/05/08", "電車",         "クライアントA社訪問",                  "東京 → 横浜（往復）",   "—",                                           1_540, "○", ""),
        (2, "2026/05/10", "タクシー（夜間）", "終電後の業務終了に伴う帰宅",            "渋谷オフィス → 自宅（世田谷区）", "—",                                "4_820", "○", "業務終了 23:45 / 終電後"),
        (3, "2026/05/14", "タクシー（昼間）", "クライアント同席の重要書類運搬",        "本社 → 大手町（B社）",  "B社 鈴木様 同乗",                            "3_180", "○", "事前申請: 2026/05/12 承認済"),
        (4, "2026/05/16", "新幹線",        "大阪支店との合同会議",                 "東京 → 新大阪（普通車指定席）", "—",                                  14_720, "○", ""),
        (5, "2026/05/16", "宿泊費",        "出張に伴う宿泊（大阪市）",              "大阪市内 ホテルABC",     "—",                                          11_500, "○", "上限13,000円以内"),
        (6, "2026/05/18", "タクシー（夜間）", "深夜帰宅（業務終了 24:30）",            "本社 → 自宅（練馬区）",   "—",                                       16_240, "○", "上限超過: F-021提出済（部長・総務部長承認）"),
        (7, "2026/05/20", "接待交際費",    "新規開拓 C社 役員との会食",             "銀座 居酒屋XYZ",        "C社 田中部長／C社 佐々木課長／当社 山田", 28_600, "○", "F-009 事前申請済 / 1名@7,150円"),
        (8, "2026/05/22", "消耗品",        "プリンタートナー購入",                 "Amazon",               "—",                                          4_180, "○", "Web明細スクショ添付"),
        (9, "2026/05/24", "電車",         "総務会議出席",                         "東京 → 品川（往復）",   "—",                                            560, "—", "3,000円未満のため明細スクショ"),
    ]

    detail_start = detail_header_row + 1
    for offset, row in enumerate(sample_rows):
        r = detail_start + offset
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "#,##0"
                if isinstance(value, str):
                    try:
                        cell.value = int(value.replace("_", "").replace(",", ""))
                    except ValueError:
                        pass
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            if offset % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)
        ws.row_dimensions[r].height = 22

    blank_rows = 12
    for offset in range(len(sample_rows), len(sample_rows) + blank_rows):
        r = detail_start + offset
        for col_idx in range(1, 10):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx == 1:
                cell.value = offset + 1
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 7:
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "#,##0"
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 22

    total_row = detail_start + len(sample_rows) + blank_rows + 1
    ws.merge_cells(f"A{total_row}:F{total_row}")
    ws[f"A{total_row}"] = "合計（税込）／Total"
    ws[f"A{total_row}"].font = _body_font(11, bold=True)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=GREY_BG)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws[f"A{total_row}"].border = BORDER_ALL

    last_detail_row = detail_start + len(sample_rows) + blank_rows - 1
    total_cell = ws.cell(row=total_row, column=7)
    total_cell.value = f"=SUM(G{detail_start}:G{last_detail_row})"
    total_cell.number_format = "¥#,##0"
    total_cell.font = _body_font(12, bold=True, color=ACCENT)
    total_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    total_cell.fill = PatternFill("solid", fgColor=GREY_BG)
    total_cell.border = BORDER_ALL

    ws.merge_cells(f"H{total_row}:I{total_row}")
    ws[f"H{total_row}"].border = BORDER_ALL
    ws[f"H{total_row}"].fill = PatternFill("solid", fgColor=GREY_BG)

    advance_row = total_row + 1
    ws.merge_cells(f"A{advance_row}:F{advance_row}")
    ws[f"A{advance_row}"] = "仮払金（既受領分）／Cash Advance Received"
    ws[f"A{advance_row}"].font = _body_font(10)
    ws[f"A{advance_row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws[f"A{advance_row}"].border = BORDER_ALL
    ws.cell(row=advance_row, column=7, value=0).number_format = "¥#,##0"
    ws.cell(row=advance_row, column=7).alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws.cell(row=advance_row, column=7).border = BORDER_ALL
    ws.merge_cells(f"H{advance_row}:I{advance_row}")
    ws[f"H{advance_row}"].border = BORDER_ALL

    net_row = advance_row + 1
    ws.merge_cells(f"A{net_row}:F{net_row}")
    ws[f"A{net_row}"] = "差引精算額（会社負担）／Net Reimbursement"
    ws[f"A{net_row}"].font = _body_font(11, bold=True)
    ws[f"A{net_row}"].fill = PatternFill("solid", fgColor=WARN_BG)
    ws[f"A{net_row}"].alignment = Alignment(horizontal="right", vertical="center", indent=1)
    ws[f"A{net_row}"].border = BORDER_ALL
    net_cell = ws.cell(row=net_row, column=7, value=f"=G{total_row}-G{advance_row}")
    net_cell.number_format = "¥#,##0"
    net_cell.font = _body_font(12, bold=True, color=ACCENT)
    net_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    net_cell.fill = PatternFill("solid", fgColor=WARN_BG)
    net_cell.border = BORDER_ALL
    ws.merge_cells(f"H{net_row}:I{net_row}")
    ws[f"H{net_row}"].border = BORDER_ALL
    ws[f"H{net_row}"].fill = PatternFill("solid", fgColor=WARN_BG)

    sig_header_row = net_row + 2
    sig_labels = ["申請者", "直属上長", "部門長", "経理部", "総務部"]
    sig_cols = ["A", "C", "E", "G", "H"]
    sig_widths = [("A", "B"), ("C", "D"), ("E", "F"), ("G", "G"), ("H", "I")]
    for label, (start_col, end_col) in zip(sig_labels, sig_widths):
        ws.merge_cells(f"{start_col}{sig_header_row}:{end_col}{sig_header_row}")
        cell = ws[f"{start_col}{sig_header_row}"]
        cell.value = label
        cell.font = _body_font(10, bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        cell.font = _title_font(10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_ALL

        ws.merge_cells(f"{start_col}{sig_header_row + 1}:{end_col}{sig_header_row + 1}")
        cell2 = ws[f"{start_col}{sig_header_row + 1}"]
        cell2.border = BORDER_ALL
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[sig_header_row + 1].height = 60

        ws.merge_cells(f"{start_col}{sig_header_row + 2}:{end_col}{sig_header_row + 2}")
        cell3 = ws[f"{start_col}{sig_header_row + 2}"]
        cell3.value = "（日付）"
        cell3.font = _body_font(8, color="808080")
        cell3.alignment = Alignment(horizontal="center", vertical="center")
        cell3.border = BORDER_ALL

    note_row = sig_header_row + 4
    ws.merge_cells(f"A{note_row}:I{note_row}")
    note_cell = ws[f"A{note_row}"]
    note_cell.value = (
        "※ 本書の提出期限は毎月25日17:00必着（規程第8条）。"
        "遅延の場合は翌々月精算となり、F-007 経費精算遅延始末書の添付が必要です。"
        "タクシー上限15,000円超過分は F-021 タクシー代超過利用理由書を添付してください（規程第5条第4項）。"
    )
    note_cell.font = _body_font(9, color="595959")
    note_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    note_cell.fill = PatternFill("solid", fgColor=WARN_BG)
    note_cell.border = BORDER_ALL
    ws.row_dimensions[note_row].height = 36

    category_dv = DataValidation(
        type="list",
        formula1='"電車,新幹線,飛行機,バス,タクシー（昼間）,タクシー（夜間）,自家用車,宿泊費,接待交際費,会議費,消耗品,通信費,その他"',
        allow_blank=True,
    )
    category_dv.error = "区分マスタから選択してください"
    category_dv.errorTitle = "区分エラー"
    ws.add_data_validation(category_dv)
    category_dv.add(f"C{detail_start}:C{last_detail_row}")

    receipt_dv = DataValidation(
        type="list",
        formula1='"○,—,電子明細,後日提出"',
        allow_blank=True,
    )
    ws.add_data_validation(receipt_dv)
    receipt_dv.add(f"H{detail_start}:H{last_detail_row}")

    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.6
    ws.page_margins.bottom = 0.6
    ws.print_title_rows = "1:9"


def build_taxi_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("タクシー利用明細")
    ws.sheet_view.showGridLines = False

    _set_col_widths(
        ws,
        {
            "A": 4,
            "B": 12,
            "C": 10,
            "D": 10,
            "E": 22,
            "F": 22,
            "G": 22,
            "H": 14,
            "I": 14,
            "J": 28,
        },
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = "タクシー利用明細（規程第5条 詳細記録）"
    ws["A1"].font = _title_font(14)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = (
        "夜間（23:00以降または終電後）／クライアント同席（事前申請必須）／不可抗力 のいずれかに該当する利用のみ会社負担。"
        "上限 15,000円/回。超過分は F-021 を提出のこと。"
    )
    ws["A2"].font = _body_font(9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws["A2"].fill = PatternFill("solid", fgColor=WARN_BG)
    ws.row_dimensions[2].height = 32

    headers = [
        "No.",
        "利用日",
        "乗車時刻",
        "降車時刻",
        "乗車地",
        "降車地",
        "利用目的",
        "金額（税込）",
        "上限超過",
        "F-021 / 事前申請 ステータス",
    ]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = _title_font(10)
        cell.fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 32

    rows = [
        (1, "2026/05/10", "23:50", "00:18", "渋谷オフィス（東京都渋谷区）", "自宅（世田谷区上馬）", "終電後帰宅（業務終了 23:45）", 4_820, "なし", "—"),
        (2, "2026/05/14", "13:20", "13:45", "本社（中央区八重洲）", "B社（千代田区大手町）", "クライアント同席・重要書類運搬", 3_180, "なし", "事前申請: 2026/05/12 承認済（承認者: 高橋部長）"),
        (3, "2026/05/18", "00:35", "01:20", "本社（中央区八重洲）", "自宅（練馬区豊玉北）", "深夜帰宅（業務終了 24:30）", 16_240, "あり", "F-021 提出済 / 部長: 5/19承認 / 総務部長: 5/20承認"),
    ]

    for offset, row in enumerate(rows):
        r = 5 + offset
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx in (1, 3, 4, 9):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "¥#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            if offset % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW)
        ws.row_dimensions[r].height = 24

    for offset in range(len(rows), len(rows) + 8):
        r = 5 + offset
        for col_idx in range(1, 11):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx == 1:
                cell.value = offset + 1
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "¥#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 22

    excess_dv = DataValidation(type="list", formula1='"なし,あり"', allow_blank=True)
    ws.add_data_validation(excess_dv)
    excess_dv.add("I5:I20")


def build_advance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("仮払い精算")
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {"A": 4, "B": 14, "C": 24, "D": 14, "E": 14, "F": 14, "G": 30})

    ws.merge_cells("A1:G1")
    ws["A1"] = "仮払い精算明細（規程第12条）"
    ws["A1"].font = _title_font(14)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    ws["A2"] = "出張等の前に仮払いを受けた場合、帰着後5営業日以内に本シートを用いて清算すること。"
    ws["A2"].font = _body_font(9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

    headers = ["No.", "仮払申請日", "目的", "仮払金額", "実費合計", "差額", "備考"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = _title_font(10)
        cell.fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    ws.row_dimensions[4].height = 24

    for r in range(5, 13):
        for col_idx in range(1, 8):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx == 1:
                cell.value = r - 4
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in (4, 5, 6):
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.number_format = "¥#,##0"
                if col_idx == 6:
                    cell.value = f"=D{r}-E{r}"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def build_category_master(wb: Workbook) -> None:
    ws = wb.create_sheet("区分マスタ")
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {"A": 22, "B": 16, "C": 50})

    ws["A1"] = "区分名"
    ws["B1"] = "勘定科目"
    ws["C1"] = "備考・規程参照"
    for col in ("A1", "B1", "C1"):
        ws[col].font = _title_font(10)
        ws[col].fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
        ws[col].border = BORDER_ALL
    ws.row_dimensions[1].height = 24

    rows = [
        ("電車",            "旅費交通費",   "規程第4条 / 経済的経路かつ最短時間"),
        ("新幹線",          "旅費交通費",   "規程第4条第2項 / 100km以上で普通車指定席"),
        ("飛行機",          "旅費交通費",   "規程第4条第3項 / 300km以上または2時間短縮"),
        ("バス",            "旅費交通費",   "規程第4条"),
        ("タクシー（昼間）", "旅費交通費",   "規程第5条第2項 / クライアント同席等の例外のみ"),
        ("タクシー（夜間）", "旅費交通費",   "規程第5条第1項 / 23:00以降または終電後"),
        ("自家用車",        "旅費交通費",   "規程第6条 / 1km=25円 / F-015 必須"),
        ("宿泊費",          "旅費交通費",   "規程第7条 / 地域別上限"),
        ("接待交際費",      "交際費",       "規程第10条 / 5,000円超は F-009"),
        ("会議費",          "会議費",       "社内会議の弁当代等"),
        ("消耗品",          "消耗品費",     "—"),
        ("通信費",          "通信費",       "—"),
        ("その他",          "雑費",         "上長承認必須"),
    ]
    for offset, row in enumerate(rows):
        r = 2 + offset
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.border = BORDER_ALL
            cell.font = _body_font(10)
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)


def build_instructions_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("記入要領")
    ws.sheet_view.showGridLines = False

    _set_col_widths(ws, {"A": 4, "B": 100})

    ws.merge_cells("A1:B1")
    ws["A1"] = "経費精算テンプレート_v3.xlsx 記入要領"
    ws["A1"].font = _title_font(14)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = "準拠規程: 株式会社デモ・ロジスティクス 旅費交通費・経費精算および労務手続き規定（DL-HR-RG-2024-007 第3版）"
    ws["A2"].font = _body_font(9, color="595959")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    sections = [
        ("【提出期限】", [
            "毎月25日 17:00 必着（規程第8条）。",
            "25日が土日祝日の場合は、直前の営業日に繰り上げ。",
            "期限を過ぎた場合は翌々月精算となり、F-007 経費精算遅延始末書の添付が必要。",
            "年度内に3回以上遅延した場合、人事考課のコンピテンシー評価で減点対象。",
        ]),
        ("【記入手順】", [
            "1. 「経費精算書」シートのヘッダー欄（氏名・社員番号・所属・対象期間等）を記入。",
            "2. 明細欄に1件1行で経費を記入。「区分」はドロップダウンから選択。",
            "3. 領収書欄は ○（原本提出）／—（添付なし）／電子明細／後日提出 から選択。",
            "4. タクシー利用は「タクシー利用明細」シートにも記入（規程第5条 詳細記録）。",
            "5. 仮払いを受けていた場合は「仮払い精算」シートに記入。",
            "6. 合計金額・差引精算額は自動計算（数式が破損しないよう、行の途中挿入は避けること）。",
            "7. 直属上長 → 部門長 → 経理 → 総務 の順で承認印を取得。",
        ]),
        ("【領収書の取扱い】", [
            "1件3,000円以上：原則として原本領収書を必須（電子レシート PDF/JPEG も可）。",
            "1件3,000円未満：利用日・店舗名・金額・利用目的が判別できる電子明細スクショで代替可。",
            "飲食を伴う経費：参加者全員の氏名・所属を必ず明記。1人あたり5,000円超の接待は F-009 事前申請必須。",
        ]),
        ("【タクシー代の重要ルール（規程第5条）】", [
            "夜間（23:00以降）または終電後の業務終了帰宅 → 会社負担可。",
            "昼間のタクシー → 原則NG。例外は ①クライアント同席+事前申請、②重量物/機密物運搬、③不可抗力。",
            "1回 15,000円（税込）が上限。超過分は F-021 タクシー代超過利用理由書を提出（部長＋総務部長の二者承認）。",
            "F-021 の提出期限は利用日の翌々営業日 17:00 まで。それ以降は全額自己負担。",
            "私的飲食後の帰宅、自宅最寄り駅までのタクシー、観光混在の移動は会社負担対象外。",
        ]),
        ("【フォーマット利用上の注意】", [
            "本テンプレートは v3 です。v2以前のフォーマットでの申請は受理されません（規程第9条第3項）。",
            "本ファイルは Google Drive 上で配布されます。原本のセル構造・数式を変更しないでください。",
            "編集する場合は、ローカルにダウンロードしてから記入し、ワークフローシステム Coconala-Flow にアップロードしてください。",
            "ファイル名は「経費精算_YYYY-MM_氏名.xlsx」の形式で保存することを推奨します。",
        ]),
        ("【お問い合わせ】", [
            "経費・支払関連: 経理部 鈴木 美咲（内線 1456 / suzuki.misaki@demo-logistics.example.co.jp）",
            "規程・労務関連: 総務部 佐藤 健一（内線 1234 / sato.kenichi@demo-logistics.example.co.jp）",
            "シフト・夜勤関連: 物流部 田中 浩二（内線 2210）",
        ]),
    ]

    row = 4
    for title, lines in sections:
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = _body_font(11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=LIGHT_NAVY)
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cell.border = BORDER_ALL
        ws.row_dimensions[row].height = 22
        row += 1

        for line in lines:
            ws.merge_cells(f"A{row}:B{row}")
            line_cell = ws[f"A{row}"]
            line_cell.value = "　" + line
            line_cell.font = _body_font(10)
            line_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            line_cell.border = BORDER_ALL
            ws.row_dimensions[row].height = 22
            row += 1

        row += 1


def build() -> Path:
    wb = Workbook()
    build_main_sheet(wb)
    build_taxi_sheet(wb)
    build_advance_sheet(wb)
    build_category_master(wb)
    build_instructions_sheet(wb)

    wb.properties.title = "経費精算テンプレート_v3"
    wb.properties.creator = "株式会社デモ・ロジスティクス 総務部"
    wb.properties.lastModifiedBy = "Demo-Logi-Bot"
    wb.properties.subject = "Expense Reimbursement Template (Dummy)"
    wb.properties.description = (
        "規程 DL-HR-RG-2024-007 準拠 / Dify x n8n ポートフォリオデモ用ダミーファイル"
    )
    wb.properties.keywords = "経費精算, 旅費交通費, タクシー, 規程, デモ"

    output_path = Path(__file__).resolve().parent.parent / OUTPUT_FILENAME
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    path = build()
    print(f"Generated: {path}")
